from openpyxl import load_workbook
from pathlib import Path
import json

BASE = Path(__file__).parent

ARQUIVO_XLSX = BASE / "dados" / "Dados.xlsx"
ARQUIVO_JS = BASE / "scripts" / "produtos.js"

wb = load_workbook(ARQUIVO_XLSX, data_only=True)
ws = wb.active

cabecalho = [str(c.value).strip() if c.value else "" for c in ws[1]]

produtos = []

for linha in ws.iter_rows(min_row=2, values_only=True):

    if linha[0] is None:
        continue

    registro = dict(zip(cabecalho, linha))

    produtos.append({
        "id": int(registro["ID"]),
        "nome": str(registro["Nome do Produto"]),
        "categoria": str(registro["Categoria"]),
        "pasta": str(registro["Pasta"]),
        "preco": float(registro["Preço (R$)"]),
        "estoque": int(registro["Estoque"]),
        "prazo": int(registro["Prazo"]),
        "disponivel": str(registro["Disponível (SIM/NÃO)"]).upper() == "SIM",
        "sobEncomenda": str(registro["Sob Encomenda (SIM/NÃO)"]).upper() == "SIM",
        "destaque": str(registro["Destaque (SIM/NÃO)"]).upper() == "SIM",
        "peso": float(registro["Peso (kg)"]),
        "largura": float(registro["Largura (cm)"]),
        "altura": float(registro["Altura (cm)"]),
        "profundidade": float(registro["Profundidade (cm)"]),
        "fotos": int(registro["Quantidade de Fotos"]),
        "descricao": str(registro["Descrição"])
    })

with open(ARQUIVO_JS, "w", encoding="utf-8") as f:

    f.write("export const produtos = ")
    json.dump(produtos, f, ensure_ascii=False, indent=4)
    f.write(";")

print(f"{len(produtos)} produtos exportados com sucesso.")
print(f"Arquivo criado: {ARQUIVO_JS}")
print("\n===================================")
print("Atualização concluída com sucesso!")
print("Agora pressione F5 no navegador.")
print("===================================")

input("\nPressione ENTER para fechar...")